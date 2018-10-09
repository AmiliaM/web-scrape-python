from openpyxl import Workbook
from bs4 import BeautifulSoup
import urllib3
import random

class Inductee:
    def __init__(self, name, year, category, members):
        self.name = name
        self.year = year
        self.members = members
        self.category = category

def get_inductees(n_inductees):
    http = urllib3.PoolManager()
    page = 0
    i_count = 0
    i_needed = random.sample(range(1, 323), n_inductees)
    i_retrieved = []

    while len(i_needed) > 0:
        url = 'https://www.rockhall.com/inductees/all?name=&field_inductee_induction_year=&field_induction_category=All&page=' + str(page)
        response = http.request("GET", url)
        soup = BeautifulSoup(response.data, features="html.parser")

        all_links = soup.find_all("a")
        for link in all_links:
            if "https://www.rockhall.com/inductees/" in link.get("href"):
                i_count += 1
                if i_count in i_needed:
                    print("Chose inductee " + link.get("href"))
                    i_needed.remove(i_count)
                    i_retrieved.append(link.get("href"))
        page += 1
    return i_retrieved

def get_inductee_data(i_url_list):
    print("Retrieving inductee data...")
    inductees = []
    http = urllib3.PoolManager()
    for url in i_url_list:
        response = http.request("GET", url)
        soup = BeautifulSoup(response.data, features="html.parser")
        name = soup.find("h1").get_text()
        year = soup.find('span', {"class" : "date-display-single"}).get_text()
        category = soup.find("div", {"class" : "inductee-category"}).get_text().strip()
        m = soup.find("ul", {"class" : 'inductee-members'})
        if m is None:
            members = 1
        else:
            members = m.get_text().count('\n')-1
        inductees.append(Inductee(name, year, category, members))
    return inductees

def inductees_to_xlsx(inductees):
    print("Converting data to xlsx...")
    wb = Workbook()
    ws = wb.active
    i_number = 1
    for o in inductees:
        ws['A'+str(i_number)] = o.name
        ws['B'+str(i_number)] = o.year
        ws['C'+str(i_number)] = o.category
        ws['D'+str(i_number)] = o.members
        i_number += 1
    wb.save("inductees.xlsx")


def main():
    urllib3.disable_warnings()
    i_list = get_inductees(20)
    inductees = get_inductee_data(i_list)
    inductees_to_xlsx(inductees)
    

if __name__ == "__main__":
    main()